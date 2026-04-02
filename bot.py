import os
import json
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

TOKEN = os.getenv("BOT_TOKEN")
FILE_NAME = "construction_progress.xlsx"
STATE_FILE = "user_state.json"

MAX_ENTRANCES = 20
MAX_FLOORS = 20

apartments = [
    "Наруж.стены кладка",
    "Внутр.стены кладка",
    "ПГП",
    "Гипс.штк",
    "Цем.штк",
    "Сшитый пол",
    "Стяжка",
    "Эл.кв+СС",
    "Окна ПВХ",
    "Окна Аллюм",
    "Двери"
]

mop = [
    "Внутр.стены кладка",
    "Коллектор кладка",
    "ВШ кладка",
    "ШТК",
    "Декоративная штк",
    "Сшитый пол",
    "Стяжка",
    "Плитка пол",
    "Плитка настенная",
    "Двери МОП",
    "Двери коллектор",
    "Ограждения ЛК",
    "Разводка ЭЛ",
    "Слаботочные сети",
    "Стояк отопление",
    "Стояк канализация",
    "Стояк ливневка",
    "Стояк вода"
]

percent_values = ["0","10","20","30","40","50","60","70","80","90","98","100"]

# Инициализация состояния
if not os.path.exists(STATE_FILE):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump({}, f)

def load_state():
    with open(STATE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_state():
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(user_data, f, ensure_ascii=False)

user_data = load_state()

start_keyboard = ReplyKeyboardMarkup(
    [["Start"]],
    resize_keyboard=True
)

def percent_keyboard():
    kb = []
    row = []
    for p in percent_values:
        row.append(p)
        if len(row) == 3:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    kb.append(["Пропустить"])
    kb.append(["⬅️ Назад"])
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def entrance_keyboard():
    keyboard = []
    for i in range(1, MAX_ENTRANCES + 1, 2):
        if i < MAX_ENTRANCES:
            keyboard.append([str(i), str(i+1)])
        else:
            keyboard.append([str(i)])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def floor_keyboard():
    keyboard = []
    for i in range(1, MAX_FLOORS + 1, 2):
        if i < MAX_FLOORS:
            keyboard.append([str(i), str(i+1)])
        else:
            keyboard.append([str(i)])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата","Время","Адрес","Подъезд","Этаж","Раздел","Пункт","Процент"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(FILE_NAME)

def save_excel(row):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append(row)
    wb.save(FILE_NAME)

def get_fill(percent):
    try:
        percent = int(percent)
    except:
        percent = 0
    if percent == 0:
        return PatternFill("solid", start_color="FF9999")
    elif percent < 100:
        return PatternFill("solid", start_color="FFF599")
    else:
        return PatternFill("solid", start_color="99FF99")

def format_sheet(ws):
    widths = {"A":16,"B":14,"C":35,"D":12,"E":12,"F":18,"G":40,"H":14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

def create_excel(data, all_floors=False):
    if all_floors:
        filename = f"{data['address'].replace(' ','_')}_полный_обход.xlsx"
        rows = data["all_rows"]
    else:
        filename = f"{data['address'].replace(' ','_')}_подъезд_{data['entrance']}_этаж_{data['floor']}.xlsx"
        rows = data["floor_rows"]

    wb = Workbook()
    ws = wb.active
    ws.append(["Дата","Время","Адрес","Подъезд","Этаж","Раздел","Пункт","Процент"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    format_sheet(ws)
    for row_num in range(2,len(rows)+2):
        ws[f"H{row_num}"].fill = get_fill(ws[f"H{row_num}"].value)
    wb.save(filename)
    return filename

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    user_data[uid] = {
        "step":"address",
        "floor_rows":[],
        "all_rows":[],
        "percents":[],
        "index":0
    }
    save_state()
    await update.message.reply_text("Введите адрес дома:", reply_markup=start_keyboard)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    text = update.message.text.strip()

    if text == "Start":
        await start(update, context)
        return

    if uid not in user_data:
        await update.message.reply_text("Нажмите Start", reply_markup=start_keyboard)
        return

    data = user_data[uid]
    step = data["step"]

    if text == "Завершить обход":
        if data.get("all_rows"):
            filename = create_excel(data, all_floors=True)
            with open(filename,"rb") as f:
                await update.message.reply_document(document=f, filename=filename)
            os.remove(filename)
        user_data[uid] = {"step":"address","floor_rows":[],"all_rows":[],"percents":[],"index":0}
        save_state()
        await update.message.reply_text("✅ Обход завершён", reply_markup=start_keyboard)
        await update.message.reply_text("Введите адрес дома:")
        return

    if step == "address":
        data["address"] = text
        data["step"] = "entrance"
        save_state()
        await update.message.reply_text("Выберите подъезд:", reply_markup=entrance_keyboard())
        return

    elif step == "entrance":
        data["entrance"] = text
        data["step"] = "floor"
        save_state()
        await update.message.reply_text("Выберите этаж:", reply_markup=floor_keyboard())
        return

    elif step == "floor":
        data["floor"] = text
        data["type"]="Кв"
        data["items"]=apartments
        data["index"]=0
        data["percents"]=[]
        data["floor_rows"]=[]
        data["step"]="percent"
        save_state()
        await update.message.reply_text(f"{data['type']}\n{data['items'][0]}:", reply_markup=percent_keyboard())
        return

    elif step == "percent":
        index = data["index"]
        current_item = data["items"][index]
        if text != "Пропустить" and text != "⬅️ Назад":
            now = datetime.now()
            row = [now.strftime("%d.%m.%Y"), now.strftime("%H:%M"), data["address"],
                   data["entrance"], data["floor"], data["type"], current_item, text]
            data["floor_rows"].append(row)
            data["all_rows"].append(row)

        if text == "⬅️ Назад":
            if index > 0:
                data["index"] -= 1
                save_state()
                await update.message.reply_text(f"{data['type']}\n{data['items'][data['index']]}:", reply_markup=percent_keyboard())
            return

        index = data["index"] + 1
        if index < len(data["items"]):
            data["index"] = index
            save_state()
            await update.message.reply_text(f"{data['type']}\n{data['items'][index]}:", reply_markup=percent_keyboard())
        else:
            if data["type"]=="Кв":
                data["type"]="МОП"
                data["items"]=mop
                data["index"]=0
                save_state()
                await update.message.reply_text(f"{data['type']}\n{data['items'][0]}:", reply_markup=percent_keyboard())
            else:
                filename = create_excel(data)
                with open(filename,"rb") as f:
                    await update.message.reply_document(document=f, filename=filename)
                os.remove(filename)
                data["step"]="entrance"
                save_state()
                await update.message.reply_text("Этаж завершён. Выберите следующий этаж или другой подъезд:", reply_markup=floor_keyboard())

# Инициализация Excel
init_excel()

# Запуск приложения
app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

print("Bot running...")
app.run_polling()
